from openpyxl import Workbook, load_workbook
import PySimpleGUI as sg

continuar = 0
tema = "TealMono"

while continuar != "n" and continuar != "nao":
    
    def Escolha():
        #layout
        sg.theme(tema)
        
        layout = [
            [sg.Text("Funcionário: "), sg.Combo(["Camila", "Thaiana", "Jose", "Moises", "Suellen"])],
            [sg.Text("Dia:"), sg.Combo(["01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "11", "12", "13", "14", "15", "16", "17", "18", "19", "20", "21", "22", "23", "24", "25", "26", "27", "28", "29", "30", "31"])],
            [sg.Button("Prosseguir", size=(20,1))]
        ]

        #janela
        janela = sg.Window("CAIXA RGI").layout(layout)

        event, values = janela.read()

        fun = values[0]
        dia = values[1]
        
        janela.Close()
        return fun, dia
    def Entrada():


        #layout
        sg.theme(tema)

        
        layout = [
            [sg.Text('Troco Inicial: ', size=(20,1)), sg.Input(size=(10,1), key="TrocoInicial")],
            [sg.Text('1 - ENTRADAS: ', size=(20,1)), sg.Input(size=(10,1), key="Entradas")],
            [sg.Text('2 - CARTÃO DE DÉBITO: ', size=(20,1)), sg.Input(size=(10,1), key="CartDEB")],
            [sg.Text('3 - DEPÓSITO BANCÁRIO: ', size=(20,1)), sg.Input(size=(10,1), key="DepoBAN")],
            [sg.Text('4 - CHEQUE PRÉ: ', size=(20,1)), sg.Input(size=(10,1), key="ChequePRE")],
            [sg.Text('5 - CHEQUE DIA: ', size=(20,1)), sg.Input(size=(10,1), key="ChequeDIA")],
            [sg.Text('6 - DEPÓSITO INTERNO: ', size=(20,1)), sg.Input(size=(10,1), key="DepositoINTERNO")],
            [sg.Text('7 - OFÍCIO ELETRÔNICO: ', size=(20,1)), sg.Input(size=(10,1), key="OficioELE"), sg.Text('Doc nº: ', size=(7,1)), sg.Input(size=(30,1), key="doc1")],
            [sg.Text('8 - E-INTIMAÇÃO: ', size=(20,1)), sg.Input(size=(10,1), key="Einti"), sg.Text('Doc nº: ', size=(7,1)), sg.Input(size=(30,1), key="doc2")], 
            [sg.Text('9 - ERRO DE FUNCIONÁRIO: ', size=(20,1)), sg.Input(size=(10,1), key="ErroFun"), sg.Text('Doc nº: ', size=(7,1)), sg.Input(size=(30,1), key="doc3")],
            [sg.Text('10 - ERRO DO SISTEMA: ', size=(20,1)), sg.Input(size=(10,1), key="ErroSis"), sg.Text('Doc nº: ', size=(7,1)), sg.Input(size=(30,1), key="doc4")],
            [sg.Text('11 - DMP: ', size=(20,1)), sg.Input(size=(10,1), key="dmp"), sg.Text('Doc nº: ', size=(7,1)), sg.Input(size=(30,1), key="doc5"), sg.Text('Empresa: ', size=(7,1)), sg.Input(size=(30,1), key="empresa")], 
            [sg.Text('12 - DÉBITO AUTORIZADO: ', size=(20,1)), sg.Input(size=(10,1), key="DebitoAuto"), sg.Text('Doc nº: ', size=(7,1)), sg.Input(size=(30,1), key="doc6"), sg.Text('Por: ', size=(7,1)), sg.Input(size=(30,1), key="por1")],
            [sg.Text('13 - NIHIL: ', size=(20,1)), sg.Input(size=(10,1), key="Nihil"), sg.Text('Doc nº: ', size=(7,1)), sg.Input(size=(30,1), key="doc7"), sg.Text('Por: ', size=(7,1)), sg.Input(size=(30,1), key="por2")],
            [sg.Text('14 - Quantidades de Transf: ', size=(20,1)), sg.Input(size=(10,1), key="TransfQtT"), sg.Text('Espécie: ', size=(10,1)), sg.Input(size=(10,1), key="TransfEspe"), sg.Text('Cheque Pré: ', size=(10,1)), sg.Input(size=(10,1), key="TransfPre"), sg.Text('Cheque Dia: ', size=(10,1)), sg.Input(size=(10,1), key="TransfDia")],
            [sg.Text('15- Caixa Final: ', size=(20,1)), sg.Input(size=(10,1), key="CaixaFinal")],
            [sg.Button("Enviar", size=(20,1))], 
        ]

        #janela
        janela = sg.Window("CAIXA RGI").layout(layout)

        event, values = janela.read()

        if values["TrocoInicial"] == "":
            values["TrocoInicial"] = "0"
        
        if values["Entradas"] == "":
            values["Entradas"] = "0"  
        
        if values["DepoBAN"] == "":
            values["DepoBAN"] = "0" 

        if values["ChequePRE"] == "":
            values["ChequePRE"] = "0"  

        if values["ChequeDIA"] == "":
            values["ChequeDIA"] = "0"  

        if values["DepositoINTERNO"] == "":
            values["DepositoINTERNO"] = "0"  

        if values["OficioELE"] == "":
            values["OficioELE"] = "0"  

        if values["Einti"] == "":
            values["Einti"] = "0"  

        if values["ErroFun"] == "":
            values["ErroFun"] = "0"  
        
        if values["ErroSis"] == "":
            values["ErroSis"] = "0" 

        if values["dmp"] == "":
            values["dmp"] = "0"  

        if values["DebitoAuto"] == "":
            values["DebitoAuto"] = "0"  

        if values["Nihil"] == "":
            values["Nihil"] = "0"  

        if values["CaixaFinal"] == "":
            values["CaixaFinal"] = "0"  

        if values["CartDEB"] == "":
            values["CartDEB"] = "0"  

        if values["TransfQtT"] == "":
            values["TransfQtT"] = "0" 

        if values["TransfEspe"] == "":
            values["TransfEspe"] = "0" 

        if values["TransfPre"] == "":
            values["TransfPre"] = "0" 

        if values["TransfDia"] == "":
            values["TransfDia"] = "0" 

        TrocoInicial = values["TrocoInicial"].replace(",", ".")
        Entradas = values["Entradas"].replace(",", ".")
        CartDEB = values["CartDEB"].replace(",", ".")
        DepoBAN = values["DepoBAN"].replace(",", ".")
        ChequePRE = values["ChequePRE"].replace(",", ".")
        ChequeDIA = values["ChequeDIA"].replace(",", ".")
        DepositoINTERNO = values["DepositoINTERNO"].replace(",", ".")
        OficioELE = values["OficioELE"].replace(",", ".")
        Einti = values["Einti"].replace(",", ".")
        ErroFun = values["ErroFun"].replace(",", ".")
        ErroSis = values["ErroSis"].replace(",", ".")
        dmp = values["dmp"].replace(",", ".") 
        DebitoAuto = values["DebitoAuto"].replace(",", ".")
        Nihil = values["Nihil"].replace(",", ".")
        CaixaFinal = values["CaixaFinal"].replace(",", ".")
        doc1 = values["doc1"]
        doc2 = values["doc2"]
        doc3 = values["doc3"]
        doc4 = values["doc4"]
        doc5 = values["doc5"]
        doc6 = values["doc6"]
        doc7 = values["doc7"]
        empresa = values["empresa"]
        por1 = values["por1"]
        por2 = values["por2"]
        TransfQtT = values["TransfQtT"].replace(",", ".")
        TransfEspe = values["TransfEspe"].replace(",", ".")
        TransfPre = values ["TransfPre"].replace(",", ".")
        TransfDia = values ["TransfDia"].replace(",", ".")
        
        janela.Close()  
        return TrocoInicial, Entradas, CartDEB, DepoBAN, ChequePRE, ChequeDIA, DepositoINTERNO, OficioELE, Einti, ErroFun, ErroSis, dmp, DebitoAuto, Nihil, CaixaFinal, doc1, doc2, doc3, doc4, doc5, doc6, doc7, empresa, por1, por2, TransfQtT, TransfEspe, TransfPre, TransfDia
    def Aviso():
        #layout
        sg.theme(tema)
        

        layout = [
            [sg.Text("{} {}".format(aviso, diferenca))],
            [sg.Button("Confirmar"), sg.Button("Refazer")],
        ]
        #janela
        janela = sg.Window("Alerta").layout(layout)
        events, values = janela.read()
        janela.Close()


        if events == "Confirmar":
            Saida()
        elif events == "Refazer":
            continuar = "s"
        return continuar
    def Saida():
        #layout
        sg.theme(tema)

        primeiro = [
            
            [sg.Text('TROCO INICIAL: R$ {}\n'.format(TrocoInicial))],
            [sg.Text('1 - ENTRADAS: R$ {}\n'.format(Entradas))],
            [sg.Text('2 - CARTÃO DE DÉBITO: R$ {} '.format(CartDEB))],
            [sg.Text('3 - DEPÓSITO BANCÁRIO: R$ {} '.format(DepoBAN))],
            [sg.Text('4 - CHEQUE PRÉ: R$ {} '.format(ChequePRE))],
            [sg.Text('5 - CHEQUE DIA: R$ {} '.format(ChequeDIA))],
            [sg.Text('6 - DEPÓSITO INTERNO: R$ {} '.format(DepositoINTERNO))],  
        ]

        segundo = [
            [sg.Text('7 - OFÍCIO ELETRÔNICO: R$ {}\n     Doc nº: {}'.format(OficioELE, doc1))],
            [sg.Text('8 - E-INTIMAÇÃO: R$ {}\n     Doc nº: {}'.format(Einti, doc2))],
            [sg.Text('9 - ERRO DE FUNCIONÁRIO: R$ {}\n     Doc nº: {}'.format(ErroFun, doc3))],
            [sg.Text('10 - ERRO DO SISTEMA: R$ {}\n     Doc nº: {}'.format(ErroSis, doc4))],
            [sg.Text('11 - DMP:  R$ {}\n     Doc nº: {}     Empresa: {}'.format(dmp, doc5, empresa))],
            [sg.Text('12 - DÉBITO AUTORIZADO: R$ {}\n      Doc nº: {}      Por: {}'.format(DebitoAuto, doc6, por1))],
            [sg.Text('13 - NIHIL: R$ {}\n     Doc nº: {}      Por: {}\n\n\n'.format(Nihil, doc7, por2))],
            [sg.Button("OK", size=(30,1))],
        ]

        terceiro = [
            
            [sg.Text('QUANTIDADES DE TRANSF: {}\n     ESPÉCIE:: R${}\n     CHEQUE PRÉ: R${}\n     CHEQUE DIA: R${}\n     TOTAL DE TRANSF: R$ {}\n'.format(TransfQtT, TransfEspe, TransfPre, TransfDia, totaltransf))],
            [sg.Text('ENTRADAS EM ESPÉCIE: R$ {}\n '.format(especie))],
            [sg.Text('EM CAIXA: R$ {} '.format(CaixaFinal))],
            [sg.Text("FECHAMENTO FINAL: R$ {} ".format(final))],
            [sg.Text('DIFERENÇA EM CAIXA: R$ {} '.format(diferenca))],
        ]

        layout = [
                
            [
                sg.Column(primeiro, size=(300,370)),
                sg.VSeparator(),
                sg.Column(segundo, size=(300,370)),
                sg.VSeparator(),
                sg.Column(terceiro, size=(300,370)),
            ],
            
        ]


        #janela
        janela = sg.Window("Funcionário: {} | Dia: {}".format(fun, dia)).layout(layout)

        event, values = janela.read()

        if event == "OK":
            continuar = "n"
        return continuar

    #chama a primeira função
    fun, dia = Escolha() 

    #Variaveis de funcionário

    if fun == "Camila":
        wb = load_workbook(filename= "CAIXA 1 - CAMILA.xlsx")
        nomearquivo = "CAIXA 1 - CAMILA.xlsx"
    elif fun == "Thaiana":
        wb = load_workbook(filename= "CAIXA 2 - THAIANA.xlsx")
        nomearquivo = "CAIXA 2 - THAIANA.xlsx"
    elif fun == "Jose":
        wb = load_workbook(filename= "CAIXA 3 - JOSE RICARDO.xlsx")
        nomearquivo = "CAIXA 3 - JOSE RICARDO.xlsx"
    elif fun == "Moises":
        wb = load_workbook(filename= "CAIXA 4 - MOISES.xlsx")
        nomearquivo = "CAIXA 4 - MOISES.xlsx"
    elif fun == "Suellen":
        wb = load_workbook(filename= "CAIXA 5 - SUELLEN.xlsx")
        nomearquivo = "CAIXA 5 - SUELLEN.xlsx"

    #Variaveis de dia

    if dia == "01":
        sh = wb["DIA 01"]
    elif dia == "02":
        sh = wb["DIA 02"]
    elif dia == "03":
        sh = wb["DIA 03"]
    elif dia == "04":
        sh = wb["DIA 04"]
    elif dia == "05":
        sh = wb["DIA 05"]
    elif dia == "06":
        sh = wb["DIA 06"]
    elif dia == "07":
        sh = wb["DIA 07"]
    elif dia == "08":
        sh = wb["DIA 08"]
    elif dia == "09":
        sh = wb["DIA 09"]
    elif dia == "10":
        sh = wb["DIA 10"]
    elif dia == "11":
        sh = wb["DIA 11"]
    elif dia == "12":
        sh = wb["DIA 12"]
    elif dia == "13":
        sh = wb["DIA 13"]
    elif dia == "14":
        sh = wb["DIA 14"]
    elif dia == "15":
        sh = wb["DIA 15"]
    elif dia == "16":
        sh = wb["DIA 16"]
    elif dia == "17":
        sh = wb["DIA 17"]
    elif dia == "18":
        sh = wb["DIA 18"]
    elif dia == "19":
        sh = wb["DIA 19"]
    elif dia == "20":
        sh = wb["DIA 20"]
    elif dia == "21":
        sh = wb["DIA 21"]
    elif dia == "22":
        sh = wb["DIA 22"]
    elif dia == "23":
        sh = wb["DIA 23"]
    elif dia == "24":
        sh = wb["DIA 24"]
    elif dia == "25":
        sh = wb["DIA 25"]
    elif dia == "26":
        sh = wb["DIA 26"]
    elif dia == "27":
        sh = wb["DIA 27"]
    elif dia == "28":
        sh = wb["DIA 28"]
    elif dia == "29":
        sh = wb["DIA 29"]
    elif dia == "30":
        sh = wb["DIA 30"]
    elif dia == "31":
        sh = wb["DIA 31"]

    #inserção de dados

    #chama a 2 função

    TrocoInicial, Entradas, CartDEB, DepoBAN, ChequePRE, ChequeDIA, DepositoINTERNO, OficioELE, Einti, ErroFun, ErroSis, dmp, DebitoAuto, Nihil, CaixaFinal, doc1, doc2, doc3, doc4, doc5, doc6, doc7, empresa, por1, por2, TransfQtT, TransfEspe, TransfPre, TransfDia= Entrada()


    sh["D5"] = float(TrocoInicial)
    sh["D7"] = float(Entradas)
    sh["D9"] = float(CartDEB)
    sh["D10"] = float(DepoBAN)
    sh["D11"] = float(ChequePRE)
    sh["D12"] = float(ChequeDIA)
    sh["D13"] = float(DepositoINTERNO)
    sh["D14"] = float(OficioELE)
    sh["D15"] = float(Einti)
    sh["D16"] = float(ErroFun)
    sh["D17"] = float(ErroSis)
    sh["D18"] = float(dmp)
    sh["D19"] =  float(DebitoAuto)
    sh["D20"] = float(Nihil)
    sh["I6"] = float(CaixaFinal)
    sh["F14"] = doc1
    sh["F15"] = doc2
    sh["F16"] = doc3
    sh["F17"] = doc4
    sh["F18"] = doc5
    sh["F19"] = doc6
    sh["F20"] = doc7
    sh["I18"] = empresa
    sh["I19"] = por1
    sh["I20"] = por2
    sh["D23"] = float(TransfQtT)
    sh["F23"] = float(TransfEspe)
    sh["I23"] = float(TransfPre)
    sh["I24"] = float(TransfDia)


    soma = float(CartDEB) +  float(DepoBAN) + float(ChequePRE) + float(ChequeDIA) + float(DepositoINTERNO) + float(OficioELE) + float(Einti) + float(ErroFun) + float(ErroSis) + float(dmp) + float(DebitoAuto) + float(Nihil)
    especie = float(Entradas) - soma - float(TransfPre) - float(TransfDia)
    final = especie - float(TransfEspe) + float(TrocoInicial)
    diferenca = float(CaixaFinal) - final
    totaltransf = float(TransfEspe) + float(TransfPre) + float(TransfDia)

    if diferenca > 0:
        aviso = "Sobrou: R$"
    else:
        aviso = "Faltou: R$"

    wb.save(filename = nomearquivo)
    
    Aviso()

